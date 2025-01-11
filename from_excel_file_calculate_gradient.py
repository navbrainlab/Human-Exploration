import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import ast
# import cv2
import os
import glob
import matplotlib.pyplot as plt
import math
# import seaborn as sns
from scipy.signal import find_peaks

subject_list = []


def find_csv_files_with_glob(path):
    # 使用**递归展开目录，查找所有.csv文件
    csv_files = glob.glob(os.path.join(path, '**', '*.csv'), recursive=True)
    return csv_files


directory_path = r'D:\human_data'
csv_paths = find_csv_files_with_glob(directory_path)

# for one_csv_path in csv_paths:
# print(csv_paths[0])
# csv = pd.read_csv(csv_paths[0])
# csv.keys()
for subject_csv in csv_paths:
    subject = subject_csv.split('\\')[3].split('_')[0]
    if subject.startswith('O') or subject.startswith('o') :
        subject = '10' + subject.split('ct')[1]
    # elif str(subject).startswith('10'):
    #     subject = 'Oct' + str(subject).split('10')[1]

    elif subject.startswith('N') or subject.startswith('n') :
        #         print(subject)
        subject = '11' + subject.split('ov')[1]
    # elif str(subject).startswith('11'):
    #     subject = 'Nov' + str(subject).split('11')[1]
    #         print(subject)

    subject_list.append(str(subject))
# subject_list = ['111801']

print(subject_list)
print(csv_paths)

p1p2_p1_thinking_time = np.empty([0, 30])
p1p2_p2_thinking_time = np.empty([0, 50])
p2_thinking_time = np.empty([0, 50])
p1p2_p1_reward = []
p1p2_p2_reward = []
p2_only_reward = []
# p1p2_subject_count = 0
all_event_p1 = []
all_event_p2 = []
all_event_p2only = []

p1_ponder_position = []
p2_ponder_position = []
p2only_ponder_position = []

p2_only = 0
p1p2 = 0
i = 0

subject_phase_to_thinking_time = {}

for csv_file in csv_paths:
    # print(csv_file)
    subject = csv_file.split("\\")[3].split("_")[0]
    # print(subject_id)
    if subject.startswith('O') or subject.startswith('o') :
        subject = '10' + subject.split('ct')[1]
    elif str(subject).startswith('10'):
        subject = '10' + str(subject).split('10')[1]

    elif subject.startswith('N') or subject.startswith('n') :
        #         print(subject)
        subject = '11' + subject.split('ov')[1]
    elif str(subject).startswith('11'):
        subject = '11' + str(subject).split('11')[1]
    # print(subject)

    csv = pd.read_csv(csv_file)
    df = ['reward', 'form1.response', 'phase', 'c4.started', 'b4.started', 'c4.stopped', 'b4.stopped']
    df_score_between_games = csv[['phase', 'reward']]
    # c1.startdded : p2开始的时间；c4.stopped : p2结束的时间
    # b1.startdded : p1开始的时间；b4.stopped : p1结束的时间
    phase = len(df_score_between_games['phase'].unique())

    if len(df_score_between_games['phase'].unique()) == 3:  # p1、p2都做的人
        #         plt.figure()
        df_score_between_games = csv[
            ['phase', 'reward', 'c4.started', 'b1.started', 'c1.started', 'b4.started', 'c4.stopped', 'b4.stopped']]
        #         p1_started_timestamp=df_score_between_games['b1.started']
        i += 1
        p1p2 += 1
        df_phase2 = df_score_between_games[df_score_between_games['phase'] == 'phase2']
        df_phase1 = df_score_between_games[df_score_between_games['phase'] == 'phase1']

        [x for x in df_score_between_games['b4.started'].tolist() if not math.isnan(x)]

        p1_started_timestamp = [x for x in df_score_between_games['b4.started'].tolist() if not math.isnan(x)]
        p1_stopped_timestamp = [x for x in df_score_between_games['b4.stopped'].tolist() if not math.isnan(x)]

        p2_started_timestamp = [x for x in df_score_between_games['c4.started'].tolist() if not math.isnan(x)]
        p2_stopped_timestamp = [x for x in df_score_between_games['c4.stopped'].tolist() if not math.isnan(x)]

        p1_timegap = [x - y for x, y in zip(p1_stopped_timestamp, p1_started_timestamp)]
        p2_timegap = [x - y for x, y in zip(p2_stopped_timestamp, p2_started_timestamp)]

        subject_p1 = subject + '_p1'
        subject_p2 = subject + '_p2'
        # print('!!!!!!!!!!!!!!!')

        subject_phase_to_thinking_time[subject_p1] = p1_timegap
        subject_phase_to_thinking_time[subject_p2] = p2_timegap
        # print('????????????????')
        # print(subject_phase_to_thinking_time)

        # p1_reward = [x for x in df_phase1['reward'].tolist() if not math.isnan(x)]
        # #         all_event_p1 = thinking_time_when_reward_change(all_event_p1,p1_reward,p1_timegap,20)
        #
        # p2_reward = [x for x in df_phase2['reward'].tolist() if not math.isnan(x)]
        # #         all_event_p2 = thinking_time_when_reward_change(all_event_p2,p2_reward,p2_timegap,20)
        #
        # p1p2_p1_reward.append(p1_reward)
        # p1p2_p2_reward.append(p2_reward)
        # print(p1_timegap)
        # print('_______________________________')

        ########################################

        filled_array = np.full(30, np.nan)  # 将原始数据填充到新数组中
        filled_array[:len(p1_timegap)] = p1_timegap
        #         print(p1_timegap)
        # this_subject_p1_ponder_position = outliner_detect(p1_timegap)
        # p1_ponder_position.append(this_subject_p1_ponder_position)

        p1_timegap = np.expand_dims(filled_array, axis=0)

        filled_array = np.full(50, np.nan)  # 将原始数据填充到新数组中
        filled_array[:len(p2_timegap)] = p2_timegap



        p2_timegap = np.expand_dims(filled_array, axis=0)



        p1p2_p1_thinking_time = np.concatenate((p1p2_p1_thinking_time, p1_timegap), axis=0)
        p1p2_p2_thinking_time = np.concatenate((p1p2_p2_thinking_time, p2_timegap), axis=0)





    elif len(df_score_between_games['phase'].unique()) == 2:  # 只做p2的人

        #         plt.figure()
        df_score_between_games = csv[['phase', 'reward', 'c1.started', 'c4.stopped', 'c4.started']]
        p2_only += 1
        df_phase2 = df_score_between_games[df_score_between_games['phase'] == 'phase2']

        p2_started_timestamp = [x for x in df_score_between_games['c4.started'].tolist() if not math.isnan(x)]
        p2_stopped_timestamp = [x for x in df_score_between_games['c4.stopped'].tolist() if not math.isnan(x)]

        #         p1_timegap = [x-y for x,y in zip(p1_stopped_timestamp,p1_started_timestamp)]
        p2_timegap = [x - y for x, y in zip(p2_stopped_timestamp, p2_started_timestamp)]

        subject_p2 = subject + '_p2only'
        subject_phase_to_thinking_time[subject_p2] = p2_timegap

        #
        p1_reward = None
        p1_timegap = None

        #         draw_thinking_time_and_goal(save_in_folder,subject_id,phase,p1_reward,p2_reward,p1_timegap,p2_timegap)

        #         overall_thinking_time_and_reward(save_in_folder,subject_id,phase,p1_reward,p2_reward,p1_timegap,p2_timegap)
        filled_array = np.full(50, np.nan)  # 将原始数据填充到新数组中
        filled_array[:len(p2_timegap)] = p2_timegap
        # this_subject_p2_ponder_position = outliner_detect(p2_timegap)
        # p2only_ponder_position.append(this_subject_p2_ponder_position)

        p2_timegap = np.expand_dims(filled_array, axis=0)
        # #         print(p1_timegap)
        #         print(p1_timegap.shape)
        # #         print(p1p2_p1_time.shape)
        p2_thinking_time = np.concatenate((p2_thinking_time, p2_timegap), axis=0)

    #         p2.append(max(df_phase2.reward))
    else:
        print(csv_file)
# # print('p1p2=',p1p2,'p2_only',p2_only)
# p1p2_p1_time
# # sns.heatmap(p1p2_p1_time,  cmap='viridis', annot=True)
# # 计算按列的平均值
# p1p2_p1_time_means = np.nanmean(p1p2_p1_time, axis=1)
# p1p2_p2_time_means = np.nanmean(p1p2_p2_time, axis=1)
# p2_time_means = np.nanmean(p2_time, axis=1)
# print(subject_phase_to_thinking_time)


def calculate_food_choice_index(subject_name, df):
    row_weight_1 = 10
    row_weight_2 = 5
    row_weight_3 = 1

    #     score_change_list = []
    # score = list(df.iloc[:, 3])
    #     print(score)
    phase = subject_name.split('_')[1]
    if phase == 'p1':
        dim = 3
        score = list(df.iloc[:, 3])
    else:
        dim = 4
        score = list(df.iloc[:, 4])
    all_dim_choice_index = {}
    all_dim_choice_gradient = {}
    score_change_list = []
    for i in range(dim):

        list_food = np.array(df.iloc[:, i].apply(lambda x: list(x.split(' ')[:3])))
        #         print(list_food)
        all_dim_choice_index[i] = []
        all_dim_choice_gradient[i] = []
        for round in range(len(list_food)):  # 这是一个维度在一轮中的排布情况
            if round == 0:
                #                 if i == 0:
                score_change = 0
                if i == 0:
                    score_change_list.append(score_change)
            else:
                before_round_this_dim_choice = list_food[round - 1]
                # try:
                score_change = score[int(round)] - score[int(round) - 1]

                if i == 0:
                    score_change_list.append(score_change)

            this_round_this_dim_choice = list_food[round]
            #             all_dim_choice_index[i] = []
            # 计算该维度中三种食物的choice index；
            choice_index = {}
            choice_gradient = {}
            all_food_count = {1: {0: [], 1: [], 2: []}, 2: {0: [], 1: [], 2: []}, 3: {0: [], 1: [], 2: []}}
            # food_index:{0、1、2行中的计数}

            for row in range(len(this_round_this_dim_choice)):
                # 游戏中的1~3行
                food_in_this_row = [int(x) for x in this_round_this_dim_choice[row]]
                for food in [1, 2, 3]:
                    all_food_count[food][row] = len([x for x in food_in_this_row if x == food])
            #                     print(this_round_this_dim_choice[row]
            for food in [1, 2, 3]: # 计算梯度
                # print(i)
                # print(round)
                # print(food)
                this_food_count = all_food_count[food]
                choice_index[food] = row_weight_1 * this_food_count[0] + row_weight_2 * this_food_count[
                    1] + row_weight_3 * this_food_count[2]
                if round == 0:
                    choice_gradient[food] = 0
                else:
                    # print(list(all_dim_choice_index.values())[-1])
                    # b = list(all_dim_choice_index.values())
                    # a = list(all_dim_choice_index.values())[0][-1]
                    # if i !=3:
                    before_food_index = list(all_dim_choice_index.values())[i][-1][food]
                    choice_index[food] = row_weight_1 * this_food_count[0] + row_weight_2 * this_food_count[
                        1] + row_weight_3 * this_food_count[2]
                    if (choice_index[food] - before_food_index) != 0:

                        choice_gradient[food] = score_change / (choice_index[food] - before_food_index)
                    else:
                        choice_gradient[food] = np.inf

                    # if i ==3:
                    #     # a = list(all_dim_choice_index.values())
                    #     before_food_index = list(all_dim_choice_index.values())[i][-1][food]
                    #     choice_index[food] = row_weight_1 * this_food_count[0] + row_weight_2 * this_food_count[
                    #         1] + row_weight_3 * this_food_count[2]
                    #     if (choice_index[food] - before_food_index) !=0:
                    #
                    #         choice_gradient[food] = score_change / (choice_index[food] - before_food_index)
                    #     else:
                    #         choice_gradient[food] = np.inf
            all_dim_choice_index[i].append(choice_index)
            all_dim_choice_gradient[i].append(choice_gradient)

    #                 print(round)
    #                 print(this_round_this_dim_choice)
    #                 print(before_round_this_dim_choice)

    # 分数变化
    return all_dim_choice_index, all_dim_choice_gradient, score_change_list


# 打开Excel文件
workbook = openpyxl.load_workbook(r'C:\Users\Windows11\Desktop\rearrange_data_thinking.xlsx')
# workbook = openpyxl.load_workbook(r'C:\Users\Windows11\Desktop\try_1.xlsx')
# 获取所有工作表的名称
sheet_names = workbook.sheetnames


# sheet_names = sheet_names[1:]
# print(f"所有工作表名称: {sheet_names}")
# # 遍历每个工作表

df_list = []

# print(subject_phase_to_thinking_time.keys())
for sheet_name in sheet_names:
    #     print(sheet_name)

    df = pd.read_excel(r'C:\Users\Windows11\Desktop\rearrange_data_thinking.xlsx', sheet_name=sheet_name,header=None)
    a = df
    all_dim_choice_index, all_dim_choice_gradient, score_change_list = calculate_food_choice_index(sheet_name, df)
    df['score change'] = score_change_list

    for dim ,v in all_dim_choice_index.items():
        for food in list(v[0].keys()):
            values_list = [d[food] for d in v if food in d]
            label = 'choice idx dim_'+str(dim+1)+ ' '+'food_'+str(food)
            df[label] = values_list



    for dim ,v in all_dim_choice_gradient.items():
        for food in list(v[0].keys()):
            values_list = [d[food] for d in v if food in d]
            label = 'grad dim_'+str(dim+1)+ ' '+'food_'+str(food)
            df[label] = values_list
            # print(df)



    df_list.append(df)


#         print(all_dim_choice_index)


# 创建一个Excel写入器对象
with pd.ExcelWriter(r'C:\Users\Windows11\Desktop\rearrange_data_thinking_250109.xlsx', engine='openpyxl') as writer:
    for csv_file, sheet_name in zip(df_list, sheet_names):
        # 读取CSV文件
        # 将DataFrame写入对应的工作表
        csv_file.to_excel(writer, sheet_name=sheet_name, index=False)

# 完成后，Excel文件将在当前目录下生成
